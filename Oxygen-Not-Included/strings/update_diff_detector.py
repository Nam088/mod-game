#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Oxygen Not Included - Auto Update Diff Detector
Tu dong phat hien chuoi moi / chuoi thay doi khi game cap nhat file strings_template.pot
"""

import sys, os, json

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    import polib
except ImportError:
    import subprocess
    subprocess.run(["pip3", "install", "polib", "--break-system-packages"], check=True)
    import polib

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


def categorize_key(key):
    if "BUILDINGS." in key or ".FACADES." in key or "STRINGS.BLUEPRINTS." in key:
        return "01. Công Trình & Ngoại Trang (Buildings & Skins)"
    elif "ELEMENTS." in key:
        return "02. Nguyên Tố & Vật Liệu (Elements)"
    elif "CREATURES." in key:
        return "03. Sinh Vật & Cây Trồng (Creatures & Plants)"
    elif "DUPLICANTS.ROLES." in key or "DUPLICANTS.CHOREGROUPS." in key or "DUPLICANTS.SKILLS." in key:
        return "04. Nghề Nghiệp & Kỹ Năng (Roles & Skills)"
    elif "DUPLICANTS.STATS." in key or "DUPLICANTS.STATUSITEMS." in key or "BUILDING.STATUSITEMS." in key:
        return "05. Chỉ Số & Trạng Thái (Vitals & Status Items)"
    elif "DUPLICANTS.TRAITS." in key or "DUPLICANTS.DISEASES." in key:
        return "06. Tính Cách & Mầm Bệnh (Traits & Diseases)"
    elif "ROOMS." in key:
        return "07. Phòng Ốc (Rooms)"
    elif "ITEMS.FOOD." in key:
        return "08. Thực Phẩm & Món Ăn (Food)"
    elif "EQUIPMENT." in key or "CLOTHING." in key or "ITEMS.INDUSTRIAL_PRODUCTS." in key:
        return "09. Trang Phục & Vật Phẩm (Suits & Items)"
    elif "ROBOTS." in key:
        return "10. Robot & Drone (Robots & Drones)"
    elif "WORLDS." in key or "ASTEROIDS." in key or "WORLD_TRAITS." in key:
        return "11. Thế Giới & Vũ Trụ (Worlds & Asteroids)"
    elif "RESEARCH." in key:
        return "12. Cây Công Nghệ (Tech Tree)"
    elif "COLONY_ACHIEVEMENTS." in key:
        return "13. Thành Tựu (Achievements)"
    elif "CODEX." in key:
        return "14. Bách Khoa Toàn Thư (Codex)"
    elif "UI." in key or "INPUT_BINDINGS." in key:
        return "15. Giao Diện & Phím Tắt (UI & Hotkeys)"
    return "16. Khác (Misc & Lore)"


def main():
    pot_path = sys.argv[1] if len(sys.argv) > 1 else "Oxygen-Not-Included/strings/strings_template.pot"
    po_path = sys.argv[2] if len(sys.argv) > 2 else "Oxygen-Not-Included/strings/strings.po"

    if not os.path.exists(pot_path):
        print(f"[X] Khong tim thay POT: {pot_path}")
        return
    if not os.path.exists(po_path):
        print(f"[X] Khong tim thay PO: {po_path}")
        return

    print("=" * 70)
    print("🚀 ONI UPDATE DIFF DETECTOR - CONG CU PHAT HIEN CAP NHAT GAME ONI")
    print("=" * 70)
    print(f"[*] File POT Moi  : {pot_path}")
    print(f"[*] File PO Dich  : {po_path}")

    pot = polib.pofile(pot_path)
    po = polib.pofile(po_path)

    po_dict = {e.msgctxt: {"msgid": e.msgid or "", "msgstr": e.msgstr or ""} for e in po if e.msgctxt}

    new_strings = []
    modified_strings = []
    unchanged = 0

    for entry in pot:
        ctx = entry.msgctxt
        if not ctx:
            continue
        en = entry.msgid or ""

        if ctx not in po_dict:
            new_strings.append({"key": ctx, "cat": categorize_key(ctx), "en_new": en, "en_old": "", "vi": "", "type": "MOI HOAN TOAN"})
        else:
            pe = po_dict[ctx]
            if pe["msgid"] != en:
                modified_strings.append({"key": ctx, "cat": categorize_key(ctx), "en_new": en, "en_old": pe["msgid"], "vi": pe["msgstr"], "type": "BI SUA GOC EN"})
            elif not pe["msgstr"].strip() and en.strip():
                new_strings.append({"key": ctx, "cat": categorize_key(ctx), "en_new": en, "en_old": en, "vi": "", "type": "CHUA DICH"})
            else:
                unchanged += 1

    total = len(new_strings) + len(modified_strings)

    print("=" * 70)
    print("BAO CAO KET QUA SO SANH:")
    print("=" * 70)
    print(f"- Tong chuoi trong POT moi         : {len(pot):,}")
    print(f"- Tong chuoi trong PO hien tai      : {len(po):,}")
    print(f"- Chuoi nguyen ven (Da dich chuan)  : {unchanged:,}")
    print(f"- Chuoi moi hoan toan can dich      : {len(new_strings):,}")
    print(f"- Chuoi bi Klei sua doi tieng Anh   : {len(modified_strings):,}")
    print(f"- TONG SO CHUOI CAN VIET HOA LAI    : {total:,}")
    print("=" * 70)

    if total == 0:
        print("\n🎉 Tuyệt vời! Không có chuỗi mới nào. Bản dịch đang đồng bộ 100% với game!")
        return

    output_dir = os.path.dirname(po_path)
    json_batch_path = os.path.join(output_dir, "new_update_batch.json")
    batch_data = {item["key"]: item.get("vi", "") for item in (new_strings + modified_strings)}
    with open(json_batch_path, "w", encoding="utf-8") as f:
        json.dump(batch_data, f, ensure_ascii=False, indent=2)
    print(f"\n[1] 📄 Đã tạo file JSON để dịch nhanh: {json_batch_path}")

    if HAS_EXCEL:
        excel_rows = [{"Trạng Thái": x["type"], "Phân Loại": x["cat"], "Key (msgctxt)": x["key"], "Tiếng Anh Mới": x["en_new"], "Tiếng Anh Cũ": x["en_old"], "Tiếng Việt Hiện Tại": x["vi"], "Bản Dịch Mới (Điền vào đây)": ""} for x in (new_strings + modified_strings)]
        dist_dir = os.path.join(os.path.dirname(output_dir), "dist")
        os.makedirs(dist_dir, exist_ok=True)
        xlsx_path = os.path.join(dist_dir, "New_Update_Changes.xlsx")
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            pd.DataFrame(excel_rows).to_excel(writer, index=False, sheet_name="Cập_Nhật_Mới")
            ws = writer.sheets["Cập_Nhật_Mới"]
            hf = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            for cell in ws[1]:
                cell.fill = hf
                cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["C"].width = 45
            ws.column_dimensions["D"].width = 50
        print(f"[2] 📊 Đã tạo file Excel so sánh trực quan: {xlsx_path}")

    print("\n" + "-" * 70)
    print("HUONG DAN CAC BUOC VIET HOA BAN CAP NHAT MOI:")
    print("  Buoc 1: Chep file strings_template.pot moi tu game vao thu muc strings/.")
    print("  Buoc 2: Chay: python3 Oxygen-Not-Included/strings/update_diff_detector.py")
    print("  Buoc 3: Dich cac chuoi trong file new_update_batch.json.")
    print("  Buoc 4: Nap ban dich vao master:")
    print("          python3 Oxygen-Not-Included/strings/apply_compact_batch.py Oxygen-Not-Included/strings/new_update_batch.json")
    print("  Buoc 5: Kiem tra va dong goi lai mod:")
    print("          python3 Oxygen-Not-Included/strings/verify_translation.py Oxygen-Not-Included/strings/strings.po")
    print("=" * 70)


if __name__ == "__main__":
    main()
